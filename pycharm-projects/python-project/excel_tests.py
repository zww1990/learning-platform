import pandas as pd
from openpyxl import Workbook


def writer_excel_1():
    file_path = r'D:\合并\demo-1.xlsx'
    data = {
        'table-1': [
            {'姓名': '张三', '性别': '男'},
            {'姓名': '李四', '性别': '女'}
        ],
        'table-2': [
            {'姓名': '张三风', '性别': '男'},
            {'姓名': '李四民', '性别': '女'}
        ]
    }
    with pd.ExcelWriter(file_path) as writer:
        for key, value in data.items():
            pd.DataFrame(value).to_excel(writer, sheet_name=key, index=False)
    print(f'写入[ {file_path} ]完毕')


def writer_excel_2():
    file_path = r'D:\合并\demo-2.xlsx'
    data = {
        'table-1': [
            ['姓名', '性别'],
            ['张三', '男'],
            ['李四', '女']
        ],
        'table-2': [
            ['姓名', '性别'],
            ['张三风', '男'],
            ['李四民', '女']
        ]
    }
    with pd.ExcelWriter(file_path) as writer:
        for key, value in data.items():
            pd.DataFrame(value).to_excel(writer, sheet_name=key, index=False, header=False)
    print(f'写入[ {file_path} ]完毕')


def writer_excel_3():
    file_path = r'D:\合并\demo-3.xlsx'
    wb = Workbook()
    ws1 = wb.create_sheet('table-1')
    print(ws1.iter_rows())
    ws2 = wb.create_sheet('table-2')
    print(ws2.iter_rows())
    ws3 = wb.create_sheet('table-3')
    print(ws3.iter_rows())
    wb.save(file_path)
    print(f'写入[ {file_path} ]完毕')


if __name__ == '__main__':
    # writer_excel_1()
    # writer_excel_2()
    writer_excel_3()
